# -*- coding: utf-8 -*-
import scrapy
from scrapy import Request
import json
import re
import openpyxl
from pprint import pprint
import os
import re
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import colors


# 国家队名字
guojiadui = ['汇金',
             '中央汇金资产管理公司',
             '中央汇金投资公司',
             '证金',
             '中国证券金融股份有限公司',
             '中证金融资产管理计划',
             '梧桐树投资平台公司',
             '北京凤山投资公司',
             '北京坤藤投资公司',
             '社保基金',
             '基本养老',
             '全国社会',
             '保障基金理事会']


# 读取需要爬取的股票名称
gupiao_list = []
with open('list.txt', 'r') as f:
    while 1:
        gupiao_name = f.readline().strip()
        if gupiao_name == '':
            break
        gupiao_list.append(gupiao_name)


# 构造股票id，股票名字，股票编号字典
with open('id.txt', 'r', encoding='gbk') as f:
    id_text = f.read()

id_dict = {}
for i in gupiao_list:
    try:
        id = re.findall('(.*?)' + i, id_text)[0].upper().strip()
        if "SZ" in id:
            bh = '0'
        else:
            bh = '1'
        id_dict[id] = [i,bh]
    except:
        pass


# 构造股票对应的url列表
strat_url = []
for i in id_dict:
    url = 'http://f10.eastmoney.com/NewFinanceAnalysis/MainTargetAjax?type=1&code={}'.format(i)
    strat_url.append(url)


wb = openpyxl.load_workbook('a.xlsx')
sh = wb['Sheet1']

class CdongfangSpider(scrapy.Spider):
    name = 'cdongfang'

    gupiao_name = gupiao_list.pop()
    allowed_domains = ['f10.eastmoney.com', '18.push2.eastmoney.com', 'push2.eastmoney.com', 'f10.eastmoney.com',
                       'f10.eastmoney.com']
    start_urls = strat_url

    def parse(self, response):
        id = re.findall('code=(\w+)',response.url)[0]
        column = sh.max_column + 2
        sh.cell(row=1, column=column - 1).value = id
        sh.cell(row=2, column=column - 1).value = id_dict[id][0]
        sh.cell(row=4, column=column - 1).value = '排名'
        sh.cell(row=4, column=column).value = '2020'
        sh.cell(row=4, column=column + 1).value = '2019'
        sh.cell(row=4, column=column + 2).value = '2018'

        data = json.loads(response.text)  # 包含了好几年的数据

        l = data[0:2] # 最新两年的所有数据

        two_year = []  # 最新两年的所需数据

        data_key = ['yyzsrtbzz','yyzsrtbzz','kfjlrtbzz','mll','jll','jqjzcsyl','zcfzl','ldbl','sdbl']
        for i in l:
            one_year = [i[key] for key in data_key]  # 一年所需的数据
            two_year.append(one_year)
        column += 1

        temp_l = [8,9,10,11,12,13,15,17,18]  # 数据所要填的行数
        for year in two_year:
            for j in range(len(temp_l)):
                sh.cell(row=temp_l[j], column=column).value = year[j]
            column += 1
        wb.save('a.xlsx')

        url = 'https://push2.eastmoney.com/api/qt/stock/get?secid={}.{}&ut=f057cbcbce2a86e2866ab8877db1d059&fields=f57,f58,f43,f169,f170,f168,f47,f48,f86,f46,f44,f60,f45,f168,f164,f50,f171&np=1&fltt=2&invt=2&cb=jQuery1800949845130501221_1589442842631&_=1589442842693'.format(
            id_dict[id][1],id[2:])   # id[2:]为没有编号的股票代码
        res = Request(url, callback=self.body_data)
        res.meta['column'] = column
        res.meta['id'] = id
        yield res

    def body_data(self, response):
        column = response.meta['column']
        id = response.meta['id']
        column -= 3

        sh.cell(row=19, column=column).value = re.findall('"f168":(.*?),', response.text)[0]
        sh.cell(row=21, column=column).value = re.findall('"f47":(.*?),', response.text)[0]
        sh.cell(row=22, column=column).value = re.findall('"f48":(.*?),', response.text)[0]
        wb.save('a.xlsx')
        url = 'http://f10.eastmoney.com/NewFinanceAnalysis/MainTargetAjax?type=0&code={}'.format(id)
        res = Request(url, self.new_head)
        res.meta['column'] = column
        res.meta['id'] = id
        yield res

    def new_head(self, response):
        column = response.meta['column']
        id = response.meta['id']
        data = json.loads(response.text)
        sh.cell(row=7, column=column).value = data[0]['yyzsrtbzz']
        sh.cell(row=8, column=column).value = data[0]['gsjlrtbzz']
        sh.cell(row=9, column=column).value = data[0]['kfjlrtbzz']
        sh.cell(row=10, column=column).value = data[0]['mll']
        sh.cell(row=11, column=column).value = data[0]['jll']
        sh.cell(row=12, column=column).value = data[0]['jqjzcsyl']
        sh.cell(row=14, column=column).value = data[0]['zcfzl']
        sh.cell(row=16, column=column).value = data[0]['ldbl']
        sh.cell(row=17, column=column).value = data[0]['sdbl']

        wb.save('a.xlsx')
        url = 'http://push2.eastmoney.com/api/qt/stock/get?ut=fa5fd1943c7b386f172d6893dbfba10b&invt=2&fltt=2&fields=f43,f57,f58,f169,f170,f46,f44,f51,f168,f47,f164,f163,f116,f60,f45,f52,f50,f48,f167,f117,f71,f161,f49,f530,f135,f136,f137,f138,f139,f141,f142,f144,f145,f147,f148,f140,f143,f146,f149,f55,f62,f162,f92,f173,f104,f105,f84,f85,f183,f184,f185,f186,f187,f188,f189,f190,f191,f192,f107,f111,f86,f177,f78,f110,f262,f263,f264,f267,f268,f250,f251,f252,f253,f254,f255,f256,f257,f258,f266,f269,f270,f271,f273,f274,f275,f127,f199,f128,f193,f196,f194,f195,f197,f80,f280,f281,f282,f284,f285,f286,f287&secid={}.{}&cb=jQuery112405506785612361488_1589457232614&_=1589457232640'.format(
            id_dict[id][1], id[2:])

        res = Request(url, self.last_data)
        res.meta['column'] = column
        res.meta['id'] = id
        yield res

    def last_data(self, response):
        column = response.meta['column']
        id = response.meta['id']
        sy4 = re.findall('"f162":(.*?),', response.text)[0]
        sj5 = re.findall('"f167":(.*?),', response.text)[0]
        sh.cell(row=5, column=column).value = sy4
        sh.cell(row=6, column=column).value = sj5
        url = 'http://f10.eastmoney.com/ShareholderResearch/ShareholderResearchAjax?code={}'.format(id)
        res = Request(url, callback=self.gudong_parse)
        res.meta['column'] = column
        res.meta['id'] = id
        yield res

    def gudong_parse(self, response):
        column = response.meta['column']
        id = response.meta['id']
        data = json.loads(response.text)
        try:
            l = [data['sdltgd'][0]['sdltgd'], data['sdgd'][0]['sdgd']]
        except:
            l = [[i for i in range(10)], data['sdgd'][0]['sdgd']]
        row = 23
        for gd in l:
            sum = 0
            sh.cell(row=row, column=column - 1).value = '排名'
            sh.cell(row=row, column=column).value = '占总流通股本持股比例'
            sh.cell(row=row, column=column + 1).value = '增减股'
            sh.cell(row=row, column=column + 2).value = '变动比例'
            sh.cell(row=row + 11, column=column - 1).value = '合计'
            try:
                for i in gd:
                    sh.cell(row=row + 1, column=column - 1).value = i['gdmc']
                    for j in guojiadui:
                        if j in i['gdmc']:
                            orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
                            sh.cell(row=row + 1, column=column - 1).fill = orange_fill
                    if '香港中央结算' in i['gdmc']:
                        grey_fill = PatternFill(fill_type='solid', fgColor="77887B")
                        sh.cell(row=row + 1, column=column - 1).fill = grey_fill
                    sh.cell(row=row + 1, column=column).value = i['zltgbcgbl']
                    sh.cell(row=row + 1, column=column + 1).value = i['zj']
                    if '-' in i['zj']:
                        ft = Font(color=colors.GREEN)
                        sh.cell(row=row + 1, column=column + 1).font = ft
                        sh.cell(row=row + 1, column=column + 2).font = ft
                    if '新' in i['zj'] or i['zj'].isdigit():
                        ft = Font(color=colors.RED)
                        sh.cell(row=row + 1, column=column + 1).font = ft
                    sh.cell(row=row + 1, column=column + 2).value = i['bdbl']
                    row += 1
                    sum = float(i['zltgbcgbl'][0:-1]) + sum
                sh.cell(row=row + 1, column=column).value = sum
                row += 2
            except:
                row += 12
                print('====================数据不全>>>',id_dict[id][0],'====================')

        wb.save('a.xlsx')
