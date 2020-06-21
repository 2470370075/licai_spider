# -*- coding: utf-8 -*-
import requests
import re
import time
import openpyxl
from datetime import datetime
import json
from pprint import pprint
from openpyxl.styles import PatternFill
from threading import Thread

url = 'http://17.push2.eastmoney.com/api/qt/clist/get?cb=jQuery112408395693008493316_1592706973904&pn=1&pz=261&po=1&np=1&ut=bd1d9ddb04089700cf9c27f6f7426281&fltt=2&invt=2&fid=f3&fs=m:90+t:3+f:!50&fields=f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f12,f13,f14,f15,f16,f17,f18,f20,f21,f23,f24,f25,f26,f22,f33,f11,f62,f128,f136,f115,f152,f124,f107,f104,f105,f140,f141,f207,f208,f209,f222&_=1592706974009'
url2 = 'http://14.push2.eastmoney.com/api/qt/clist/get?cb=jQuery1124062466096800215_1592718855265&pn=1&pz=61&po=1&np=1&ut=bd1d9ddb04089700cf9c27f6f7426281&fltt=2&invt=2&fid=f3&fs=m:90+t:2+f:!50&fields=f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f12,f13,f14,f15,f16,f17,f18,f20,f21,f23,f24,f25,f26,f22,f33,f11,f62,f128,f136,f115,f152,f124,f107,f104,f105,f140,f141,f207,f208,f209,f222&_=1592718855272'

wb = openpyxl.load_workbook('板块记录表1.xlsx')

def run(url,sheet):
    res = requests.get(url)
    title = re.findall('"f14":"(.*?)","f15"',res.text,)
    up_down = re.findall('"f3":(.*?),"f4"',res.text,)
    main = re.findall('"f128":"(.*?)","f140"',res.text,)

    row = 1
    column = 2

    sh = wb[sheet]
    while 1:
        res = sh.cell(row=row, column=column).value
        if isinstance(res,datetime):
            if datetime.now().date() == res.date():
                break
            else:
                column += 3

        elif res != None:
            column += 1
        else:
            break

    color_dict = {}

    for i in range(2,262):
        color = sh.cell(row=i, column=1).fill.start_color
        if color.rgb != '00000000':
            value = sh.cell(row=i, column=1).value.strip()
            color_dict[value]=color

    sh.cell(row=1, column=column).value = datetime.now().date()
    for row in range(2,len(title)+2):
        res = sh.cell(row=row, column=column)
        res.value = title.pop(0)
        sh.cell(row=row, column=column + 1).value = up_down.pop(0)
        sh.cell(row=row, column=column + 2).value = main.pop(0)


    for row in range(1,sh.max_row + 1):
        for column in range(1,sh.max_column + 1):
            res = sh.cell(row=row, column=column)
            if res.value in color_dict:
                fill = PatternFill(fill_type='solid', fgColor=color_dict[res.value])
                res.fill = fill



start_time = time.time()
#
p1 = Thread(target=run, args=(url,'概念板块',))
p1.start()
p2 = Thread(target=run, args=(url2,'行业板块',))
p2.start()
p1.join()
p2.join()

# run(url,'概念板块')
# run(url2,'行业板块')

wb.save('板块记录表1.xlsx')

end_time = time.time()
print(end_time-start_time)